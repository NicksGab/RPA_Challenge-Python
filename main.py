from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium import webdriver
import openpyxl
import time


url = 'https://rpachallenge.com/'
service = Service('chromedriver.exe')
options = webdriver.ChromeOptions()
options.add_argument("--start-maximized")
options.add_argument("--disable-automation")
options.add_experimental_option("detach", True)

driver = webdriver.Edge(service=service, options=options)
driver.get(url)





challenge = openpyxl.load_workbook('challenge.xlsx')
challenge_data = challenge.active


data_table = []

for index in range(2,12):
    first_name = challenge_data.cell(row=index, column=1).value
    last_name = challenge_data.cell(row=index, column=2).value
    company_name = challenge_data.cell(row=index, column=3).value
    role_in_company = challenge_data.cell(row=index, column=4).value
    address = challenge_data.cell(row=index, column=5).value
    email = challenge_data.cell(row=index, column=6).value
    phone = challenge_data.cell(row=index, column=7).value

    data_table.append([first_name, last_name, company_name, role_in_company, address, email, phone])



script = f'''
    let data_table = {data_table}

    let btn_start = "//button[contains(text(),'Start')]";
    document.evaluate(btn_start, document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue.click();
    
    for (let i = 0; i < data_table.length; i++) {{
    let row = data_table[i];

    function preencher(){{ 
    let input_fname = document.querySelector("input[ng-reflect-name='labelFirstName']");
    let input_lname = document.querySelector("input[ng-reflect-name='labelLastName']");
    let input_Cname = document.querySelector("input[ng-reflect-name='labelCompanyName']");
    let input_Role = document.querySelector("input[ng-reflect-name='labelRole']");
    let input_Ads = document.querySelector("input[ng-reflect-name='labelAddress']");
    let input_mail = document.querySelector("input[ng-reflect-name='labelEmail']");
    let input_fone = document.querySelector("input[ng-reflect-name='labelPhone']");
    
    input_fname.value = row[0]
    input_lname.value = row[1]
    input_Cname.value = row[2]
    input_Role.value = row[3]
    input_Ads.value = row[4]
    input_mail.value = row[5]
    input_fone.value = row[6]

    document.querySelector("input[type='submit']").click()
 }}; preencher() }};'''
    
driver.execute_script(script)




time.sleep(10)

driver.quit()